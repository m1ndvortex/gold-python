"""
Report Scheduler and Export Service for Advanced Analytics & Business Intelligence

This service provides:
- Automated report scheduling with cron-like configuration
- Multi-format export service (PDF, Excel, CSV)
- Email delivery system for scheduled reports
- Report template management and versioning
"""

from typing import Dict, List, Any, Optional, Union
from datetime import datetime, date, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, desc
import pandas as pd
import json
import uuid
import asyncio
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import io
import logging
from pathlib import Path
import schedule
import time
from threading import Thread
from concurrent.futures import ThreadPoolExecutor

# Import for PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Import for Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from database import get_db
import models
from schemas import User
from .report_engine_service import ReportEngineService

logger = logging.getLogger(__name__)

class ReportSchedulerService:
    """
    Report scheduling and export service
    """
    
    def __init__(self, db_session: Session):
        self.db = db_session
        self.report_engine = ReportEngineService(db_session)
        self.export_formats = ['pdf', 'excel', 'csv', 'json']
        self.scheduler_running = False
        self.scheduled_jobs = {}
        
        # Email configuration
        self.smtp_server = os.getenv('SMTP_SERVER', 'localhost')
        self.smtp_port = int(os.getenv('SMTP_PORT', '587'))
        self.smtp_username = os.getenv('SMTP_USERNAME', '')
        self.smtp_password = os.getenv('SMTP_PASSWORD', '')
        self.smtp_use_tls = os.getenv('SMTP_USE_TLS', 'true').lower() == 'true'
        
        # Export directory
        self.export_dir = Path(os.getenv('REPORT_EXPORT_DIR', '/tmp/exports'))
        self.export_dir.mkdir(parents=True, exist_ok=True)

    async def schedule_report(
        self, 
        report_config: Dict[str, Any], 
        schedule_config: Dict[str, Any],
        recipients: List[str],
        created_by: str
    ) -> Dict[str, Any]:
        """
        Schedule a report for automated generation and delivery
        
        Args:
            report_config: Report configuration
            schedule_config: Schedule configuration (cron-like)
            recipients: List of email addresses
            created_by: User ID who created the schedule
            
        Returns:
            Dictionary with schedule details
        """
        try:
            logger.info(f"Scheduling report: {report_config.get('name', 'Unnamed')}")
            
            # Validate schedule configuration
            self._validate_schedule_config(schedule_config)
            
            # Create scheduled report record
            scheduled_report = models.ScheduledReport(
                id=uuid.uuid4(),
                name=report_config.get('name', 'Scheduled Report'),
                description=report_config.get('description', ''),
                report_config=report_config,
                schedule_config=schedule_config,
                recipients=recipients,
                export_formats=schedule_config.get('export_formats', ['pdf']),
                is_active=True,
                created_by=uuid.UUID(created_by),
                created_at=datetime.now(),
                next_run_at=self._calculate_next_run(schedule_config)
            )
            
            self.db.add(scheduled_report)
            self.db.commit()
            
            # Add to scheduler
            self._add_to_scheduler(scheduled_report)
            
            return {
                'schedule_id': str(scheduled_report.id),
                'name': scheduled_report.name,
                'next_run_at': scheduled_report.next_run_at.isoformat(),
                'recipients': recipients,
                'export_formats': scheduled_report.export_formats,
                'status': 'scheduled'
            }
            
        except Exception as e:
            logger.error(f"Error scheduling report: {str(e)}")
            raise Exception(f"Failed to schedule report: {str(e)}")

    def _validate_schedule_config(self, config: Dict[str, Any]) -> None:
        """Validate schedule configuration"""
        required_fields = ['frequency']
        
        for field in required_fields:
            if field not in config:
                raise ValueError(f"Missing required schedule field: {field}")
        
        frequency = config.get('frequency')
        valid_frequencies = ['daily', 'weekly', 'monthly', 'custom']
        
        if frequency not in valid_frequencies:
            raise ValueError(f"Invalid frequency: {frequency}. Must be one of {valid_frequencies}")
        
        if frequency == 'custom' and 'cron_expression' not in config:
            raise ValueError("Custom frequency requires cron_expression")

    def _calculate_next_run(self, schedule_config: Dict[str, Any]) -> datetime:
        """Calculate next run time based on schedule configuration"""
        frequency = schedule_config.get('frequency')
        time_of_day = schedule_config.get('time', '09:00')
        
        # Parse time
        try:
            hour, minute = map(int, time_of_day.split(':'))
        except ValueError:
            hour, minute = 9, 0
        
        now = datetime.now()
        
        if frequency == 'daily':
            next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(days=1)
        
        elif frequency == 'weekly':
            day_of_week = schedule_config.get('day_of_week', 1)  # Monday = 1
            days_ahead = day_of_week - now.weekday()
            if days_ahead <= 0:  # Target day already happened this week
                days_ahead += 7
            next_run = now + timedelta(days=days_ahead)
            next_run = next_run.replace(hour=hour, minute=minute, second=0, microsecond=0)
        
        elif frequency == 'monthly':
            day_of_month = schedule_config.get('day_of_month', 1)
            next_run = now.replace(day=day_of_month, hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                # Move to next month
                if next_run.month == 12:
                    next_run = next_run.replace(year=next_run.year + 1, month=1)
                else:
                    next_run = next_run.replace(month=next_run.month + 1)
        
        else:  # custom
            # For custom cron expressions, calculate next run
            # This is a simplified implementation - in production, use a proper cron parser
            next_run = now + timedelta(hours=1)  # Default to 1 hour from now
        
        return next_run

    async def export_report(
        self, 
        report_data: Dict[str, Any], 
        export_format: str,
        filename: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Export report data to specified format
        
        Args:
            report_data: Report data from ReportEngineService
            export_format: Export format (pdf, excel, csv, json)
            filename: Optional custom filename
            
        Returns:
            Dictionary with export details
        """
        try:
            if export_format not in self.export_formats:
                raise ValueError(f"Unsupported export format: {export_format}")
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                report_name = report_data.get('name', 'report').replace(' ', '_')
                filename = f"{report_name}_{timestamp}.{export_format}"
            
            filepath = self.export_dir / filename
            
            # Export based on format
            if export_format == 'csv':
                export_result = await self._export_to_csv(report_data, filepath)
            elif export_format == 'excel':
                export_result = await self._export_to_excel(report_data, filepath)
            elif export_format == 'pdf':
                export_result = await self._export_to_pdf(report_data, filepath)
            elif export_format == 'json':
                export_result = await self._export_to_json(report_data, filepath)
            else:
                raise ValueError(f"Export format {export_format} not implemented")
            
            return {
                'filename': filename,
                'filepath': str(filepath),
                'format': export_format,
                'size_bytes': filepath.stat().st_size if filepath.exists() else 0,
                'exported_at': datetime.now().isoformat(),
                **export_result
            }
            
        except Exception as e:
            logger.error(f"Error exporting report: {str(e)}")
            raise Exception(f"Failed to export report: {str(e)}")

    async def _export_to_csv(self, report_data: Dict[str, Any], filepath: Path) -> Dict[str, Any]:
        """Export report to CSV format"""
        data = report_data.get('data', [])
        
        if not data:
            # Create empty CSV with headers
            df = pd.DataFrame()
        else:
            df = pd.DataFrame(data)
        
        df.to_csv(filepath, index=False)
        
        return {
            'rows_exported': len(data),
            'columns_exported': len(df.columns) if not df.empty else 0
        }

    async def _export_to_excel(self, report_data: Dict[str, Any], filepath: Path) -> Dict[str, Any]:
        """Export report to Excel format"""
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl library not available for Excel export")
        
        data = report_data.get('data', [])
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        if data:
            df = pd.DataFrame(data)
            
            # Add headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary = report_data.get('summary', {})
        
        summary_ws.cell(row=1, column=1, value="Report Summary").font = Font(bold=True, size=14)
        row = 3
        
        for key, value in summary.items():
            summary_ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
            summary_ws.cell(row=row, column=2, value=str(value))
            row += 1
        
        wb.save(filepath)
        
        return {
            'rows_exported': len(data),
            'columns_exported': len(df.columns) if data else 0,
            'sheets_created': len(wb.sheetnames)
        }

    async def _export_to_pdf(self, report_data: Dict[str, Any], filepath: Path) -> Dict[str, Any]:
        """Export report to PDF format"""
        if not REPORTLAB_AVAILABLE:
            raise Exception("reportlab library not available for PDF export")
        
        data = report_data.get('data', [])
        summary = report_data.get('summary', {})
        
        # Create PDF document
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph(report_data.get('name', 'Report'), title_style))
        story.append(Spacer(1, 12))
        
        # Summary section
        if summary:
            story.append(Paragraph("Summary", styles['Heading2']))
            summary_data = []
            for key, value in summary.items():
                summary_data.append([key.replace('_', ' ').title(), str(value)])
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 20))
        
        # Data section
        if data:
            story.append(Paragraph("Data", styles['Heading2']))
            
            # Prepare table data
            if data:
                headers = list(data[0].keys())
                table_data = [headers]
                
                for row in data[:100]:  # Limit to first 100 rows for PDF
                    table_data.append([str(row.get(header, '')) for header in headers])
                
                # Create table
                data_table = Table(table_data)
                data_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8)
                ]))
                story.append(data_table)
                
                if len(data) > 100:
                    story.append(Spacer(1, 12))
                    story.append(Paragraph(f"Note: Showing first 100 rows of {len(data)} total rows", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        return {
            'rows_exported': min(len(data), 100),
            'total_rows': len(data),
            'pages_created': 1  # Simplified - would need page counting logic
        }

    async def _export_to_json(self, report_data: Dict[str, Any], filepath: Path) -> Dict[str, Any]:
        """Export report to JSON format"""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, default=str)
        
        return {
            'rows_exported': len(report_data.get('data', [])),
            'file_size_bytes': filepath.stat().st_size
        }

    async def send_report_email(
        self, 
        recipients: List[str], 
        report_name: str,
        export_files: List[Dict[str, Any]],
        custom_message: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Send report via email with attachments
        
        Args:
            recipients: List of email addresses
            report_name: Name of the report
            export_files: List of exported file details
            custom_message: Optional custom message
            
        Returns:
            Dictionary with send results
        """
        try:
            logger.info(f"Sending report email to {len(recipients)} recipients")
            
            # Create message
            msg = MIMEMultipart()
            msg['From'] = self.smtp_username
            msg['To'] = ', '.join(recipients)
            msg['Subject'] = f"Scheduled Report: {report_name}"
            
            # Email body
            body = custom_message or f"""
            Dear Recipient,
            
            Please find attached the scheduled report: {report_name}
            
            Report Details:
            - Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            - Formats included: {', '.join([f['format'] for f in export_files])}
            
            Best regards,
            Gold Shop Analytics System
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach files
            for file_info in export_files:
                filepath = Path(file_info['filepath'])
                if filepath.exists():
                    with open(filepath, 'rb') as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                    
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {file_info["filename"]}'
                    )
                    msg.attach(part)
            
            # Send email
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            if self.smtp_use_tls:
                server.starttls()
            
            if self.smtp_username and self.smtp_password:
                server.login(self.smtp_username, self.smtp_password)
            
            server.sendmail(self.smtp_username, recipients, msg.as_string())
            server.quit()
            
            return {
                'sent': True,
                'recipients': recipients,
                'attachments': len(export_files),
                'sent_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error sending report email: {str(e)}")
            return {
                'sent': False,
                'error': str(e),
                'recipients': recipients
            }

    def _add_to_scheduler(self, scheduled_report) -> None:
        """Add scheduled report to the scheduler"""
        schedule_id = str(scheduled_report.id)
        
        # Remove existing schedule if it exists
        if schedule_id in self.scheduled_jobs:
            schedule.cancel_job(self.scheduled_jobs[schedule_id])
        
        # Add new schedule
        frequency = scheduled_report.schedule_config.get('frequency')
        time_of_day = scheduled_report.schedule_config.get('time', '09:00')
        
        if frequency == 'daily':
            job = schedule.every().day.at(time_of_day).do(
                self._execute_scheduled_report, scheduled_report.id
            )
        elif frequency == 'weekly':
            day_of_week = scheduled_report.schedule_config.get('day_of_week', 1)
            day_names = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
            day_name = day_names[day_of_week - 1]
            job = getattr(schedule.every(), day_name).at(time_of_day).do(
                self._execute_scheduled_report, scheduled_report.id
            )
        elif frequency == 'monthly':
            # For monthly, we'll check daily and execute on the right day
            job = schedule.every().day.at(time_of_day).do(
                self._check_monthly_schedule, scheduled_report.id
            )
        
        self.scheduled_jobs[schedule_id] = job

    async def _execute_scheduled_report(self, schedule_id: uuid.UUID) -> None:
        """Execute a scheduled report"""
        try:
            # Get scheduled report
            scheduled_report = self.db.query(models.ScheduledReport).filter(
                models.ScheduledReport.id == schedule_id,
                models.ScheduledReport.is_active == True
            ).first()
            
            if not scheduled_report:
                logger.warning(f"Scheduled report {schedule_id} not found or inactive")
                return
            
            logger.info(f"Executing scheduled report: {scheduled_report.name}")
            
            # Generate report
            report_data = await self.report_engine.build_custom_report(
                scheduled_report.report_config
            )
            
            # Export in requested formats
            export_files = []
            for export_format in scheduled_report.export_formats:
                export_result = await self.export_report(
                    report_data, 
                    export_format,
                    f"{scheduled_report.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{export_format}"
                )
                export_files.append(export_result)
            
            # Send email
            email_result = await self.send_report_email(
                scheduled_report.recipients,
                scheduled_report.name,
                export_files
            )
            
            # Update last run time and next run time
            scheduled_report.last_run_at = datetime.now()
            scheduled_report.next_run_at = self._calculate_next_run(scheduled_report.schedule_config)
            scheduled_report.run_count = (scheduled_report.run_count or 0) + 1
            
            if email_result.get('sent'):
                scheduled_report.last_success_at = datetime.now()
            else:
                scheduled_report.last_error = email_result.get('error', 'Unknown error')
            
            self.db.commit()
            
            logger.info(f"Scheduled report {scheduled_report.name} executed successfully")
            
        except Exception as e:
            logger.error(f"Error executing scheduled report {schedule_id}: {str(e)}")
            
            # Update error information
            try:
                scheduled_report = self.db.query(models.ScheduledReport).filter(
                    models.ScheduledReport.id == schedule_id
                ).first()
                if scheduled_report:
                    scheduled_report.last_error = str(e)
                    scheduled_report.error_count = (scheduled_report.error_count or 0) + 1
                    self.db.commit()
            except:
                pass

    def _check_monthly_schedule(self, schedule_id: uuid.UUID) -> None:
        """Check if monthly scheduled report should run today"""
        try:
            scheduled_report = self.db.query(models.ScheduledReport).filter(
                models.ScheduledReport.id == schedule_id,
                models.ScheduledReport.is_active == True
            ).first()
            
            if not scheduled_report:
                return
            
            day_of_month = scheduled_report.schedule_config.get('day_of_month', 1)
            if datetime.now().day == day_of_month:
                asyncio.create_task(self._execute_scheduled_report(schedule_id))
                
        except Exception as e:
            logger.error(f"Error checking monthly schedule {schedule_id}: {str(e)}")

    def start_scheduler(self) -> None:
        """Start the report scheduler"""
        if self.scheduler_running:
            return
        
        self.scheduler_running = True
        
        def run_scheduler():
            while self.scheduler_running:
                schedule.run_pending()
                time.sleep(60)  # Check every minute
        
        scheduler_thread = Thread(target=run_scheduler, daemon=True)
        scheduler_thread.start()
        
        logger.info("Report scheduler started")

    def stop_scheduler(self) -> None:
        """Stop the report scheduler"""
        self.scheduler_running = False
        schedule.clear()
        self.scheduled_jobs.clear()
        logger.info("Report scheduler stopped")

    async def get_scheduled_reports(self, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """Get list of scheduled reports"""
        query = self.db.query(models.ScheduledReport)
        
        if user_id:
            query = query.filter(models.ScheduledReport.created_by == uuid.UUID(user_id))
        
        scheduled_reports = query.order_by(desc(models.ScheduledReport.created_at)).all()
        
        return [
            {
                'id': str(report.id),
                'name': report.name,
                'description': report.description,
                'is_active': report.is_active,
                'export_formats': report.export_formats,
                'recipients': report.recipients,
                'schedule_config': report.schedule_config,
                'next_run_at': report.next_run_at.isoformat() if report.next_run_at else None,
                'last_run_at': report.last_run_at.isoformat() if report.last_run_at else None,
                'last_success_at': report.last_success_at.isoformat() if report.last_success_at else None,
                'run_count': report.run_count or 0,
                'error_count': report.error_count or 0,
                'last_error': report.last_error,
                'created_at': report.created_at.isoformat()
            }
            for report in scheduled_reports
        ]